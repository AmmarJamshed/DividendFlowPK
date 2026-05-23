#!/usr/bin/env python3
"""Generate deployment calendar CSV + Excel workbook in docs/."""
from __future__ import annotations

import csv
import os
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
DOCS.mkdir(parents=True, exist_ok=True)

ROWS = [
    (
        "1",
        "dividendflow-backend",
        "Continuous",
        "Runs 24/7 after deploy",
        "Render Web",
        "REST API; serves data CSVs; AI and market endpoints",
    ),
    (
        "2",
        "dividendflow-frontend",
        "Continuous",
        "Runs 24/7 after deploy",
        "Render Static",
        "Serves built React app",
    ),
    (
        "3",
        "Backend build",
        "On demand",
        "Each backend service deploy",
        "Render",
        "npm install in backend; node server.js",
    ),
    (
        "4",
        "Frontend build",
        "On demand",
        "Each frontend service deploy",
        "Render",
        "npm install && npm run build; publish build/",
    ),
    (
        "5",
        "dividendflow-scraper",
        "Scheduled",
        "Daily 11:00 UTC → 16:00 PKT",
        "Render Cron",
        "run-all.js — dividend merge, CSVs, optional GitHub push + email",
    ),
    (
        "6",
        "dividendflow-news",
        "Scheduled",
        "Daily 12:00 UTC → 17:00 PKT",
        "Render Cron",
        "run-news.js — news, prices, Groq, GitHub push",
    ),
    (
        "7",
        "dividendflow-health-check",
        "Scheduled",
        "Every 6 h: 00:00 06:00 12:00 18:00 UTC → 05/11/17/23 PKT",
        "Render Cron",
        "health-check.js pings BACKEND_URL",
    ),
    (
        "9",
        "PSX Market Closing Prices (GitHub Actions)",
        "Scheduled",
        "Daily 12:00 UTC → 17:00 PKT",
        "GitHub Actions",
        "psx.py — prices + full payouts CSVs; commit/push",
    ),
    (
        "10",
        "PSX Market Closing Prices",
        "On demand",
        "Actions → Run workflow",
        "GitHub Actions",
        "Same as row 9 (manual)",
    ),
    (
        "11",
        "deploy-render.ps1",
        "On demand",
        "Whenever you run the script",
        "Local + Render API",
        "Triggers deploy for all Blueprint services",
    ),
    (
        "12",
        "Git push → Render",
        "On demand",
        "Push to connected branch (e.g. main)",
        "GitHub + Render",
        "Auto build & deploy if enabled",
    ),
    (
        "13",
        "Render Manual Deploy",
        "On demand",
        "Dashboard → Manual Deploy",
        "Render UI",
        "Per-service redeploy",
    ),
]

HEADERS = [
    "#",
    "Item",
    "Schedule or demand",
    "When / detail (UTC → PKT)",
    "Platform",
    "Description",
]

SUMMARY_ROWS = [
    ("Continuous", "2", "dividendflow-backend; dividendflow-frontend"),
    (
        "Scheduled",
        "4",
        "dividendflow-scraper; dividendflow-news; dividendflow-health-check; GitHub PSX cron",
    ),
    (
        "On demand",
        "6",
        "Backend/frontend builds (per deploy); PSX manual workflow; deploy-render.ps1; "
        "git push; Render manual deploy",
    ),
]

SUMMARY_HEADERS = ["Schedule or demand", "Count", "Items"]

CRON_ROWS = [
    ("dividendflow-scraper", "0 11 * * *", "Daily 11:00 UTC"),
    ("dividendflow-news", "0 12 * * *", "Daily 12:00 UTC"),
    ("dividendflow-health-check", "0 */6 * * *", "Every 6 hours on the hour"),
    ("GitHub: PSX Market Closing Prices", "0 12 * * *", "Daily 12:00 UTC"),
]

CRON_HEADERS = ["Service / workflow", "Cron (UTC)", "Human readable"]


def write_csv_utf8_bom(path: Path, headers: list[str], rows: list[tuple]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        w.writerow(headers)
        w.writerows(rows)


def write_xlsx(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D9488")  # teal-600
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, ncol: int) -> None:
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sheet 1 — All items
    ws1 = wb.active
    ws1.title = "All items"
    ws1.append(HEADERS)
    for r in ROWS:
        ws1.append(list(r))
    style_header(ws1, len(HEADERS))
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = wrap
    widths = [4, 38, 18, 42, 22, 52]
    for i, w in enumerate(widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws1.max_row}"

    # Sheet 2 — Summary
    ws2 = wb.create_sheet("Summary by type")
    ws2.append(SUMMARY_HEADERS)
    for r in SUMMARY_ROWS:
        ws2.append(list(r))
    style_header(ws2, len(SUMMARY_HEADERS))
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.alignment = wrap
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 70

    # Sheet 3 — Cron reference
    ws3 = wb.create_sheet("Cron reference")
    ws3.append(CRON_HEADERS)
    for r in CRON_ROWS:
        ws3.append(list(r))
    style_header(ws3, len(CRON_HEADERS))
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 28

    wb.save(path)


def main() -> None:
    csv_main = DOCS / "deployment-calendar.csv"
    csv_summary = DOCS / "deployment-calendar-summary.csv"
    csv_cron = DOCS / "deployment-calendar-cron.csv"
    xlsx_path = DOCS / "deployment-calendar.xlsx"

    write_csv_utf8_bom(csv_main, HEADERS, ROWS)
    write_csv_utf8_bom(csv_summary, SUMMARY_HEADERS, SUMMARY_ROWS)
    write_csv_utf8_bom(csv_cron, CRON_HEADERS, CRON_ROWS)
    print(f"Wrote {csv_main}")
    print(f"Wrote {csv_summary}")
    print(f"Wrote {csv_cron}")

    try:
        write_xlsx(xlsx_path)
        print(f"Wrote {xlsx_path}")
    except ImportError:
        print("Skipping .xlsx (install: pip install openpyxl)")


if __name__ == "__main__":
    main()
